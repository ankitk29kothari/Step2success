import pandas as pd
import time
import os
import openpyxl

#another library for excel which work on cell basis not like pandas whole sheet basis
file_name='input.xlsx'


#open excel	
workbook_obj = openpyxl.load_workbook(file_name)
#open sheet no   
sheet_obj = workbook_obj['Sheet1']
    

#writting a value to specific cell   
sheet_obj.cell(row=5, column=4).value = "Hi, Your great"

# reading a specific cell
read_data=(sheet_obj.cell(row=4, column=1).value)
    
print(read_data)
workbook_obj.save(file_name)






#opening the file at end of the program
os.system(file_name)

