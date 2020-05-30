import concurrent.futures
from concurrent.futures import ProcessPoolExecutor
import time
import paramiko
import time
import pandas as pd
import time
import openpyxl
import sys
import threading
host='10.57.164.20'

username='nirai0'
password='Nishtharai15@'
gsk_login='hhj70755'
gsk_password='India@2wsx'

excel_name="input.xlsx"
global_lock = threading.Lock()

def read_config():
	


def UpdateRoutersResultsToSheet(s_id,fping,output):

	print('writting')
	return()
	while global_lock.locked():
		time.sleep(0.01)
		continue
	global_lock.acquire()
	workbook_obj = openpyxl.load_workbook(excel_name)
	
	sheet_obj = workbook_obj['Sheet1']
	
	sheet_obj.cell(row=s_id, column=3).value = fping
	sheet_obj.cell(row=s_id, column=4).value = output
	workbook_obj.save(excel_name)
	print('Excel save',s_id)
	global_lock.release()

def UpdateRoutersResultsToSheet_fping(s_id,fping,output):
	
	#while global_lock.locked():
	#	time.sleep(0.01)
	#	continue
	global_lock.acquire()
	workbook_obj = openpyxl.load_workbook(excel_name)
	
	sheet_obj = workbook_obj['Sheet1']

	result=str(output).strip()
	input1=result.split('input rate')[-1]
	input1=input1.split(",")[0]
	output1=result.split('output rate')[-1]
	output1=output1.split(",")[0]
	input1=input1.replace('bits/sec','')
	output1=output1.replace('bits/sec','')

	
	sheet_obj.cell(row=s_id, column=3).value = fping
	sheet_obj.cell(row=s_id, column=4).value = input1
	sheet_obj.cell(row=s_id, column=5).value = output1
	sheet_obj.cell(row=s_id, column=11).value = result
	workbook_obj.save(excel_name)
	print('Excel save',s_id,input1,output1)
	global_lock.release()

def ssh_connect():
	
	global client1
	client1 = paramiko.SSHClient()
		
	client1.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client1.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)

	global client2
	client2 = paramiko.SSHClient()
		
	client2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client2.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)


	global client3
	client3 = paramiko.SSHClient()
		
	client3.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client3.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)


	global client4
	client4 = paramiko.SSHClient()
		
	client4.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client4.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)
	print('4th')

	global client5
	client5 = paramiko.SSHClient()
		
	client5.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client5.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)

	global client6
	client6 = paramiko.SSHClient()
		
	client6.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client6.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)
	print('6th')

	global client7
	client7 = paramiko.SSHClient()
		
	client7.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client7.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)


	global client8
	client8 = paramiko.SSHClient()
		
	client8.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client8.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)
	print('8th')


	global client9
	client9 = paramiko.SSHClient()
		
	client9.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client9.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)


	global client10
	client10 = paramiko.SSHClient()
		
	client10.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	
	client10.connect(hostname=host, username=username, password=password, look_for_keys=False, allow_agent=False)


def do_something(s_id,device,command):
	try:
	
		if (s_id%10==0):
			my_client=client10
		elif (s_id%9==0):
			my_client=client9
		elif (s_id%8==0):
			my_client=client8
		elif (s_id%7==0):
			my_client=client7
		elif (s_id%6==0):
			my_client=client6
		elif (s_id%5==0):
			my_client=client5
		elif (s_id%4==0):
			my_client=client4
		elif(s_id%3==0):
			my_client=client3
		elif(s_id%2==0):
			my_client=client2
		else:
			my_client=client1
		print('start',s_id,device,command)

		
		command22 = "/usr/local/sbin/fping {0}".format(device)
		stdin, stdout, stderr = my_client.exec_command(command22, bufsize=1024)
		error = stderr.read()
		data = stdout.read()
		if error:
			error=str(error.decode("utf-8"))
			print(error)
			UpdateRoutersResultsToSheet_fping(s_id,error,'')
			return()
		if data:
			data=str(data.decode("utf-8"))
			


		shell_object = my_client.invoke_shell()
		time.sleep (0.2)



		output = str(shell_object.recv(999999))
		
		while not 'TACACS Nmstel password' in output:
			time.sleep(0.1)
			#print(".")
			output = str(shell_object.recv(999999))
			
		shell_object.send(password + "\n")
		#print('Shell Invoked')


		time.sleep (0.2)
		output = str(shell_object.recv(999999))
		while not 'Nmstel credentials registered successfully' in output:
			time.sleep(0.1)
			#print('.')
			output = str(shell_object.recv(999999))
			

		#print('Tacas Registered successfully')
		
		shell_object.send("l {}".format(device)+'\n')
		time.sleep(1)

		
		output = str(shell_object.recv(999999))
		c=0
		while not 'Device connected' in output:
			c+=1
			#print(c)
			if 'Please enter login credentials' in output:
				print("Entering Gsk logins")
				shell_object.send("{}".format(gsk_login)+'\n')
				time.sleep(3)
				shell_object.send("{}".format(gsk_password)+'\n')
				time.sleep(3)

			time.sleep(0.2)
			#print('..')
			output = str(shell_object.recv(999999))
			#print(output)
			if c>10:
				print(output)
				return("error")

			continue
		#print('Device connected firing command')
		time.sleep(1.5)
		shell_object.send(command+'\n')
		time.sleep(3)
		output = (shell_object.recv(999999))
		output=str(output.decode("utf-8"))
	
		UpdateRoutersResultsToSheet_fping(s_id,"reachable",output)
		
		return(output)
	except Exception as e:
		print(e)


def ConnectControlNetAndRunCommand(device):
    pass
    try:
      text = []
      
    
      current_index=(devices.index(device))+2
      ssh = paramiko.SSHClient()
      ssh.load_system_host_keys()
      ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
      time.sleep(0.01)
      
      
      ssh.connect(hostname=host, port=22, username=username, password=password)
     

      command22 = "/usr/local/sbin/fping {0}".format(device)
      stdin, stdout, stderr = ssh.exec_command(command22, bufsize=1024)
      error = stderr.read()
      data = stdout.read()
      if error:
        decode_data=(error.decode("utf-8"))
        UpdateRoutersResultsToSheet_fping(device,decode_data,current_index) 
      if data:
        decode_data=(data.decode("utf-8"))
        UpdateRoutersResultsToSheet_fping(device,decode_data,current_index)
        return(data)
    except Exception as e:
      print("connection failed\n", e)


def main():
	df = pd.read_excel(excel_name)
	#df=df[df.fping.isnull()]
	print(len(df))
	global devices
	devices=list(df['switch'])
	s_ids=list(df['id'])
	commands=list(df['command'])
	print('main')
	
	t1=time.perf_counter()
	ssh_connect()

	#with concurrent.futures.ProcessPoolExecutor() as executor:
	with concurrent.futures.ThreadPoolExecutor(max_workers = 40) as executor:
		

	

		#a=executor.submit(do_something,1)
		#print(a.result())
		results=executor.map(do_something,s_ids ,devices,commands)

		try:
			for result in results:
				print(result)

				#print(result)
		except Exception as e:
			print(e)
		
		
		t2=time.perf_counter()
		print('Finsihed in',(t2-t1/2),'sec')




if __name__ == '__main__':
	main()